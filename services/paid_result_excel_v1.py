"""services/paid_result_excel_v1.py — Premium Excel export v1

INPUT  = premium_result_v1 public projection only
OUTPUT = XLSX bytes

formatter 이지 판정기가 아니다.
DB 0 · HTTP 0 · LEG 0 · LLM 0 · datetime.now 0 · random 0.
법령 재판정 0 · 정렬/renumber/dedupe 0 · 원문 trim/정규화 0.
"""
from __future__ import annotations

import io
import json
import re
from typing import Any, Dict, Iterable, List, Sequence, Tuple
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

EXCEL_CELL_MAX = 32767
CANONICAL_UNAVAILABLE = "법령 원문 확인 필요"
SHEET_NAMES: Tuple[str, ...] = (
    "Summary",
    "Obligations",
    "Schedule",
    "Assignments",
    "Evidence",
)
SUMMARY_FIELDS: Tuple[Tuple[str, Tuple[str, ...]], ...] = (
    ("diagnosed_at", ("diagnosis", "diagnosed_at")),
    ("company_name", ("profile", "company_name")),
    ("sector", ("profile", "sector")),
    ("workers", ("profile", "workers")),
    ("floor_area", ("profile", "floor_area")),
    ("contract_amount_eok", ("profile", "contract_amount_eok")),
    ("construction_type", ("profile", "construction_type")),
    ("building_use_type", ("profile", "building_use_type")),
    ("address", ("profile", "address")),
    ("has_excavation", ("profile", "has_excavation")),
    ("has_hazardous_material", ("profile", "has_hazardous_material")),
    ("total_obligation_count", ("materials", "overview", "total_obligation_count")),
    ("distinct_law_count", ("materials", "overview", "distinct_law_count")),
)
OBLIGATION_COLUMNS: Tuple[str, ...] = (
    "ref",
    "law_name",
    "law_article",
    "content_type",
    "obligation_type",
    "who",
    "recipient",
    "condition",
    "where",
    "how",
    "when",
    "inspection_cycle",
    "check_result",
    "canonical_source_text",
)
SCHEDULE_COLUMNS: Tuple[str, ...] = (
    "ref",
    "law_name",
    "law_article",
    "when",
    "inspection_cycle",
    "raw_cycle",
    "conflict",
)
ASSIGNMENT_COLUMNS: Tuple[str, ...] = (
    "ref",
    "law_name",
    "law_article",
    "who",
    "recipient",
    "where",
    "how",
    "condition",
)
EVIDENCE_COLUMNS: Tuple[str, ...] = (
    "law_name",
    "article_no",
    "article_sub_no",
    "article_title",
    "article_text",
    "related_refs_json",
    "article_part_index",
    "article_part_count",
    "article_text_part",
)
_HEADER_FONT = Font(bold=True)
_ALIGN = Alignment(wrap_text=True, vertical="top")
_FILENAME_SAFE = re.compile(r'[\\/:*?"<>|\x00-\x1f]+')


def _as_dict(value: Any) -> Dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _as_list(value: Any) -> List[Any]:
    return list(value) if isinstance(value, list) else []


def _dig(src: Any, path: Sequence[str]) -> Any:
    cur = src
    for key in path:
        if not isinstance(cur, dict):
            return None
        cur = cur.get(key)
    return cur


def _present(value: Any) -> bool:
    return value is not None and value != ""


def split_excel_text(text: str, limit: int = EXCEL_CELL_MAX) -> List[str]:
    """Deterministic split. concat(parts) == original. 부분삭제 0."""
    if text == "":
        return [""]
    return [text[i:i + limit] for i in range(0, len(text), limit)]


def canonical_source_text(premium_result: Dict[str, Any], obligation_ref: Any) -> str:
    """EXACT-ONE fail-closed. first-wins 0 · dedupe 0. legal.evidence 대체 0."""
    matches: List[Any] = []
    for item in _as_list(premium_result.get("canonical_sources")):
        if not isinstance(item, dict):
            continue
        if item.get("ref") != obligation_ref:
            continue
        text = item.get("text")
        if not isinstance(text, str) or text == "":
            continue
        matches.append(text)
    if len(matches) == 1:
        return matches[0]
    return CANONICAL_UNAVAILABLE


def related_refs_json(related_refs: Any) -> str:
    refs = list(related_refs) if isinstance(related_refs, list) else []
    return json.dumps(refs, ensure_ascii=False, separators=(",", ":"))


def excel_filename(premium_result: Dict[str, Any]) -> str:
    """stored 값만. datetime.now 0."""
    company = _dig(premium_result, ("profile", "company_name"))
    diagnosed = _dig(premium_result, ("diagnosis", "diagnosed_at"))
    company_s = str(company).strip() if _present(company) else ""
    diagnosed_s = str(diagnosed).strip() if _present(diagnosed) else ""
    if not company_s:
        return "TAI_법령진단_결과.xlsx"
    cname = _FILENAME_SAFE.sub("_", company_s).strip(" ._") or "결과"
    if not diagnosed_s:
        return "TAI_법령진단_{}.xlsx".format(cname)
    stamp = _FILENAME_SAFE.sub("_", diagnosed_s).strip(" ._")
    if not stamp:
        return "TAI_법령진단_{}.xlsx".format(cname)
    return "TAI_법령진단_{}_{}.xlsx".format(cname, stamp)


def content_disposition(filename: str) -> str:
    ascii_fallback = "TAI_diagnosis.xlsx"
    return "attachment; filename=\"{}\"; filename*=UTF-8''{}".format(
        ascii_fallback, quote(filename, safe=""),
    )


def _set_cell(cell, value: Any) -> None:
    """문자열은 항상 string type. =/+/@/- 시작이어도 formula 실행 0. apostrophe 삽입 0."""
    if value is None:
        cell.value = None
        return
    if isinstance(value, bool):
        cell.value = value
        return
    if isinstance(value, int):
        cell.value = value
        return
    if isinstance(value, float):
        cell.value = value
        return
    if not isinstance(value, str):
        value = str(value)
    cell.value = value
    cell.data_type = "s"


def _write_header(ws: Worksheet, columns: Sequence[str]) -> None:
    for col, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGN
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}1".format(get_column_letter(len(columns)))
    for col, name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col)].width = min(max(len(name) + 4, 14), 48)
    ws.row_dimensions[1].height = 18


def _write_row(ws: Worksheet, row_idx: int, values: Sequence[Any]) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col)
        _set_cell(cell, value)
        cell.alignment = _ALIGN


def _style_used(ws: Worksheet) -> None:
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = "A1:{}{}".format(
            get_column_letter(ws.max_column), ws.max_row,
        )


def _write_summary(ws: Worksheet, premium: Dict[str, Any]) -> None:
    _write_header(ws, ("field", "value"))
    row = 2
    for field, path in SUMMARY_FIELDS:
        value = _dig(premium, path)
        if not _present(value):
            continue
        _write_row(ws, row, (field, value))
        row += 1
    _style_used(ws)


def _obligation_base(ob: Dict[str, Any]) -> Dict[str, Any]:
    legal = _as_dict(ob.get("legal"))
    classification = _as_dict(ob.get("classification"))
    duty = _as_dict(ob.get("duty"))
    timing = _as_dict(ob.get("timing"))
    applicability = _as_dict(ob.get("applicability"))
    verification = _as_dict(ob.get("verification"))
    return {
        "ref": ob.get("ref"),
        "law_name": legal.get("law_name"),
        "law_article": legal.get("law_article"),
        "content_type": classification.get("content_type"),
        "obligation_type": classification.get("obligation_type"),
        "who": duty.get("who"),
        "recipient": duty.get("recipient"),
        "condition": applicability.get("condition"),
        "where": duty.get("where"),
        "how": duty.get("how"),
        "when": timing.get("when"),
        "inspection_cycle": timing.get("inspection_cycle"),
        "raw_cycle": timing.get("raw_cycle"),
        "conflict": timing.get("conflict"),
        "check_result": verification.get("check_result"),
    }


def _write_obligations(ws: Worksheet, premium: Dict[str, Any], rows: List[Dict[str, Any]]) -> None:
    _write_header(ws, OBLIGATION_COLUMNS)
    for i, base in enumerate(rows, start=2):
        text = canonical_source_text(premium, base.get("ref"))
        _write_row(ws, i, [base.get(col) if col != "canonical_source_text" else text for col in OBLIGATION_COLUMNS])
    _style_used(ws)


def _write_named(ws: Worksheet, columns: Sequence[str], rows: List[Dict[str, Any]]) -> None:
    _write_header(ws, columns)
    for i, base in enumerate(rows, start=2):
        _write_row(ws, i, [base.get(col) for col in columns])
    _style_used(ws)


def _write_evidence(ws: Worksheet, articles: Iterable[Any]) -> None:
    _write_header(ws, EVIDENCE_COLUMNS)
    excel_row = 2
    for article in articles:
        if not isinstance(article, dict):
            continue
        original = article.get("article_text")
        if original is None:
            original = ""
        if not isinstance(original, str):
            original = str(original)
        parts = split_excel_text(original)
        part_count = len(parts)
        refs_json = related_refs_json(article.get("related_refs"))
        for idx, part in enumerate(parts, start=1):
            article_text_cell = original if part_count == 1 else None
            _write_row(ws, excel_row, [
                article.get("law_name"),
                article.get("article_no"),
                article.get("article_sub_no"),
                article.get("article_title"),
                article_text_cell,
                refs_json,
                idx,
                part_count,
                part,
            ])
            excel_row += 1
    _style_used(ws)


def build_paid_result_excel_v1(premium_result: dict) -> bytes:
    """premium_result_v1 → XLSX bytes. 입력 mutation 0."""
    premium = _as_dict(premium_result)
    obligations = [
        _obligation_base(ob)
        for ob in _as_list(_as_dict(premium.get("materials")).get("obligations"))
        if isinstance(ob, dict)
    ]
    wb = Workbook()
    ws0 = wb.active
    ws0.title = SHEET_NAMES[0]
    _write_summary(ws0, premium)
    ws1 = wb.create_sheet(SHEET_NAMES[1])
    _write_obligations(ws1, premium, obligations)
    ws2 = wb.create_sheet(SHEET_NAMES[2])
    _write_named(ws2, SCHEDULE_COLUMNS, obligations)
    ws3 = wb.create_sheet(SHEET_NAMES[3])
    _write_named(ws3, ASSIGNMENT_COLUMNS, obligations)
    ws4 = wb.create_sheet(SHEET_NAMES[4])
    _write_evidence(ws4, _as_list(_as_dict(premium.get("evidence")).get("articles")))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = [
    "CANONICAL_UNAVAILABLE",
    "EXCEL_CELL_MAX",
    "SHEET_NAMES",
    "build_paid_result_excel_v1",
    "canonical_source_text",
    "content_disposition",
    "excel_filename",
    "related_refs_json",
    "split_excel_text",
]
