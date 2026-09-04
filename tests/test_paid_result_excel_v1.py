"""tests/test_paid_result_excel_v1.py — E01~E15 + E06-B

INPUT = premium_result_v1 only. 실 LEG/DB 0. binary xlsx repo commit 0.
"""
from __future__ import annotations

import hashlib
import io
import json
import os
import asyncio

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

import routers.diagnosis_result_web as rw
from services.paid_result_excel_v1 import (
    CANONICAL_UNAVAILABLE,
    EXCEL_CELL_MAX,
    SHEET_NAMES,
    build_paid_result_excel_v1,
    content_disposition,
    excel_filename,
)
from tests.test_paid_result_delivery_wiring_v1 import (
    install,
    leg_obligation,
    source_item,
    stored_rec,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FORBIDDEN = (
    "atom_id", "source_atom_ids", "source_index", "source_sha256",
    "factory_id", "company_id", "payment_ref", "ci_hash",
    "raw_structured_input", "provenance", "identity",
)
SECRET_WHAT = "DUTY_WHAT_MUST_NOT_LEAK"
SECRET_EVIDENCE = "LEGAL_EVIDENCE_MUST_NOT_BE_CANONICAL"


def _ob(ref, *, article="1", law="산업안전보건법", when="상시", cycle=None,
        raw="상시", who="사업주", condition=None, where=None, how=None,
        ctype="OBLIGATION", otype="ACTION", recipient=None, check="VERIFIED",
        conflict=False, evidence=SECRET_EVIDENCE):
    return {
        "ref": ref,
        "legal": {"law_name": law, "law_article": article, "evidence": evidence},
        "classification": {"content_type": ctype, "obligation_type": otype},
        "duty": {"who": who, "recipient": recipient, "where": where, "how": how, "what": SECRET_WHAT},
        "applicability": {"condition": condition},
        "verification": {"check_result": check},
        "timing": {"when": when, "inspection_cycle": cycle, "raw_cycle": raw, "conflict": conflict},
        "decision_input_count": 1,
    }


def _premium(**overrides):
    base = {
        "version": 1,
        "contract_version": 1,
        "diagnosis": {"diagnosed_at": "2026-08-11T01:25:02+00:00"},
        "profile": {
            "profile_version": 1,
            "company_name": "샘플 건설현장",
            "sector": "CONSTRUCTION",
            "workers": 86,
            "floor_area": None,
            "contract_amount_eok": 53,
            "construction_type": "건축",
            "building_use_type": None,
            "address": "경기도 화성시",
            "has_excavation": True,
            "has_hazardous_material": False,
        },
        "materials": {
            "overview": {"total_obligation_count": 2, "distinct_law_count": 1},
            "obligations": [
                _ob(17, article="375", how="=1+1"),
                _ob(18, article="375", how="+cmd"),
            ],
        },
        "evidence": {
            "articles": [
                {
                    "law_name": "산업안전보건기준에 관한 규칙",
                    "article_no": 375,
                    "article_sub_no": None,
                    "article_title": "화물취급 작업",
                    "article_text": "제375조 원문 그대로",
                    "related_refs": [17, 18],
                },
                {
                    "law_name": "산업안전보건기준에 관한 규칙",
                    "article_no": 221,
                    "article_sub_no": 2,
                    "article_title": "충돌위험 방지조치",
                    "article_text": "제221조의2 원문",
                    "related_refs": [4],
                },
                {
                    "law_name": "산업안전보건기준에 관한 규칙",
                    "article_no": 221,
                    "article_sub_no": 3,
                    "article_title": "좌석안전띠의 착용",
                    "article_text": "제221조의3 원문",
                    "related_refs": [6],
                },
                {
                    "law_name": "산업안전보건기준에 관한 규칙",
                    "article_no": 221,
                    "article_sub_no": 4,
                    "article_title": "잠금장치의 체결",
                    "article_text": "제221조의4 원문",
                    "related_refs": [5],
                },
                {
                    "law_name": "산업안전보건기준에 관한 규칙",
                    "article_no": 221,
                    "article_sub_no": 5,
                    "article_title": "인양작업 시 조치",
                    "article_text": "제221조의5 원문",
                    "related_refs": [3],
                },
            ]
        },
        "canonical_sources": [
            {"ref": 17, "text": "원문-17-EXACT"},
            {"ref": 18, "text": "원문-18-EXACT"},
        ],
    }
    base.update(overrides)
    return base


def _load(raw: bytes):
    return load_workbook(io.BytesIO(raw), data_only=False)


def _rows(ws, skip_header=True):
    values = list(ws.iter_rows(values_only=True))
    return values[1:] if skip_header and values else values


def _header(ws):
    return [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]


def _cell_values(wb):
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                out.append(cell.value)
    return out


def test_E01_workbook_valid():
    raw = build_paid_result_excel_v1(_premium())
    wb = _load(raw)
    assert wb is not None
    wb.close()


def test_E02_five_sheets_exact():
    wb = _load(build_paid_result_excel_v1(_premium()))
    assert tuple(wb.sheetnames) == SHEET_NAMES
    wb.close()


def test_E03_obligation_cardinality():
    p = _premium()
    obs = p["materials"]["obligations"]
    p["materials"]["overview"]["total_obligation_count"] = len(obs)
    wb = _load(build_paid_result_excel_v1(p))
    body = _rows(wb["Obligations"])
    assert len(body) == p["materials"]["overview"]["total_obligation_count"] == len(obs)
    wb.close()


def test_E04_ref_preservation_no_renumber():
    p = _premium()
    p["materials"]["obligations"] = [_ob(18), _ob(17), _ob(4)]
    p["materials"]["overview"]["total_obligation_count"] = 3
    wb = _load(build_paid_result_excel_v1(p))
    refs = [r[0] for r in _rows(wb["Obligations"])]
    assert refs == [18, 17, 4]
    wb.close()


def test_E05_canonical_exact_bytes():
    p = _premium()
    wb = _load(build_paid_result_excel_v1(p))
    header = _header(wb["Obligations"])
    col = header.index("canonical_source_text")
    rows = _rows(wb["Obligations"])
    by_ref = {r[0]: r[col] for r in rows}
    assert by_ref[17] == "원문-17-EXACT"
    assert by_ref[18] == "원문-18-EXACT"
    wb.close()


def test_E06_canonical_fail_closed_no_fallback():
    p = _premium()
    p["canonical_sources"] = [{"ref": 99, "text": "다른원문"}]
    wb = _load(build_paid_result_excel_v1(p))
    header = _header(wb["Obligations"])
    col = header.index("canonical_source_text")
    texts = [r[col] for r in _rows(wb["Obligations"])]
    assert texts == [CANONICAL_UNAVAILABLE, CANONICAL_UNAVAILABLE]
    dumped = "\n".join("" if v is None else str(v) for v in _cell_values(wb))
    assert SECRET_WHAT not in dumped
    assert SECRET_EVIDENCE not in dumped
    wb.close()


def test_E06B_duplicate_canonical_fail_closed_even_if_identical():
    p = _premium()
    p["materials"]["obligations"] = [_ob(17)]
    p["materials"]["overview"]["total_obligation_count"] = 1
    p["canonical_sources"] = [
        {"ref": 17, "text": "A"},
        {"ref": 17, "text": "A"},
    ]
    wb = _load(build_paid_result_excel_v1(p))
    header = _header(wb["Obligations"])
    col = header.index("canonical_source_text")
    assert _rows(wb["Obligations"])[0][col] == CANONICAL_UNAVAILABLE
    wb.close()


def test_E07_evidence_article_text_exact():
    p = _premium()
    original = p["evidence"]["articles"][0]["article_text"]
    wb = _load(build_paid_result_excel_v1(p))
    header = _header(wb["Evidence"])
    text_col = header.index("article_text")
    title_col = header.index("article_title")
    found = None
    for r in _rows(wb["Evidence"]):
        if r[title_col] == "화물취급 작업":
            found = r[text_col]
            break
    assert found == original
    wb.close()


def test_E08_related_refs_article375_two_to_two():
    p = _premium()
    wb = _load(build_paid_result_excel_v1(p))
    eh = _header(wb["Evidence"])
    refs_col = eh.index("related_refs_json")
    no_col = eh.index("article_no")
    evidence_refs = None
    for r in _rows(wb["Evidence"]):
        if r[no_col] == 375:
            evidence_refs = r[refs_col]
            break
    assert evidence_refs == "[17,18]"
    oh = _header(wb["Obligations"])
    ref_col = oh.index("ref")
    art_col = oh.index("law_article")
    ob_rows = [r for r in _rows(wb["Obligations"]) if r[art_col] == "375"]
    refs = [r[ref_col] for r in ob_rows]
    assert refs == [17, 18]
    assert len(refs) == 2
    wb.close()


def test_E09_subarticle_221_kept_distinct():
    wb = _load(build_paid_result_excel_v1(_premium()))
    h = _header(wb["Evidence"])
    no_col = h.index("article_no")
    sub_col = h.index("article_sub_no")
    subs = [r[sub_col] for r in _rows(wb["Evidence"]) if r[no_col] == 221]
    assert subs == [2, 3, 4, 5]
    wb.close()


def test_E10_no_duty_what_in_code_or_result():
    src_path = os.path.join(os.path.dirname(__file__), "..", "services", "paid_result_excel_v1.py")
    src = open(src_path, encoding="utf-8").read()
    assert '.get("what")' not in src
    assert "['what']" not in src
    assert '["what"]' not in src
    wb = _load(build_paid_result_excel_v1(_premium()))
    dumped = "\n".join("" if v is None else str(v) for v in _cell_values(wb))
    assert SECRET_WHAT not in dumped
    assert "what" not in _header(wb["Obligations"])
    wb.close()


def test_E11_forbidden_internals_absent():
    wb = _load(build_paid_result_excel_v1(_premium()))
    dumped = "\n".join("" if v is None else str(v) for v in _cell_values(wb))
    for key in FORBIDDEN:
        assert key not in dumped, key
        for ws in wb.worksheets:
            assert key not in _header(ws)
    wb.close()
    src_path = os.path.join(os.path.dirname(__file__), "..", "services", "paid_result_excel_v1.py")
    src = open(src_path, encoding="utf-8").read()
    for key in FORBIDDEN:
        assert key not in src, key


def test_E12_formula_safety_not_formula_cells():
    p = _premium()
    p["profile"]["company_name"] = "=1+1"
    p["profile"]["address"] = "@foo"
    p["materials"]["obligations"] = [
        _ob(1, how="=1+1"),
        _ob(2, who="+cmd"),
        _ob(3, where="@foo"),
        _ob(4, condition="-1+1"),
    ]
    p["canonical_sources"] = [
        {"ref": 1, "text": "=1+1"},
        {"ref": 2, "text": "+cmd"},
        {"ref": 3, "text": "@foo"},
        {"ref": 4, "text": "-1+1"},
    ]
    raw = build_paid_result_excel_v1(p)
    wb = _load(raw)
    formulas = []
    restored = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    formulas.append((ws.title, cell.coordinate, cell.value))
                if cell.value in ("=1+1", "+cmd", "@foo", "-1+1"):
                    restored.append(cell.value)
                    assert cell.data_type == "s"
    assert formulas == []
    assert "=1+1" in restored
    assert "+cmd" in restored
    assert "@foo" in restored
    wb.close()


def test_E13_long_text_no_truncation_concat_equals_original():
    original = ("가나다라마바사아자차" * 4000) + "끝마커"
    assert len(original) > EXCEL_CELL_MAX
    p = _premium()
    p["evidence"]["articles"] = [{
        "law_name": "테스트법",
        "article_no": 1,
        "article_sub_no": None,
        "article_title": "장문",
        "article_text": original,
        "related_refs": [1],
    }]
    wb = _load(build_paid_result_excel_v1(p))
    h = _header(wb["Evidence"])
    part_col = h.index("article_text_part")
    idx_col = h.index("article_part_index")
    cnt_col = h.index("article_part_count")
    text_col = h.index("article_text")
    parts = _rows(wb["Evidence"])
    assert parts
    assert all(r[cnt_col] == len(parts) for r in parts)
    assert [r[idx_col] for r in parts] == list(range(1, len(parts) + 1))
    concat = "".join(r[part_col] or "" for r in parts)
    assert concat == original
    for r in parts:
        chunk = r[part_col] or ""
        assert len(chunk) <= EXCEL_CELL_MAX
        full = r[text_col]
        if full is not None:
            assert len(str(full)) <= EXCEL_CELL_MAX
    wb.close()


def test_E14_route_mime_and_disposition(monkeypatch):
    rec = stored_rec([leg_obligation("a0", "점검")], tier="BUILDING_V2")
    install(monkeypatch, rec, product_items=[source_item(0, "a0", "원문A")])
    resp = rw.get_paid_result_excel("tok-1")
    assert resp.media_type == XLSX_MIME
    disp = resp.headers["Content-Disposition"]
    assert disp.startswith("attachment;")
    assert "filename" in disp
    chunks = []

    async def _drain():
        async for chunk in resp.body_iterator:
            chunks.append(chunk if isinstance(chunk, (bytes, bytearray)) else bytes(chunk))

    asyncio.run(_drain())
    raw = b"".join(chunks)
    wb = _load(raw)
    assert tuple(wb.sheetnames) == SHEET_NAMES
    wb.close()


def test_E15_paid_access_parity(monkeypatch):
    rec = stored_rec([leg_obligation("a0", "점검")], tier="BUILDING_V2")
    install(monkeypatch, rec, product_items=[source_item(0, "a0", "원문A")])
    json_ok = rw.get_paid_result_web("tok-1")
    assert json_ok["data"]["is_free"] is False
    assert "premium_result_v1" in json_ok["data"]
    xlsx = rw.get_paid_result_excel("tok-1")
    assert xlsx.media_type == XLSX_MIME

    install(monkeypatch, None)
    with pytest.raises(HTTPException) as not_found_json:
        rw.get_paid_result_web("missing")
    with pytest.raises(HTTPException) as not_found_xlsx:
        rw.get_paid_result_excel("missing")
    assert not_found_json.value.status_code == 404
    assert not_found_xlsx.value.status_code == 404

    rec_inactive = stored_rec([leg_obligation("a0", "점검")], tier="BUILDING_V2", status="INACTIVE")
    install(monkeypatch, rec_inactive, product_items=[source_item(0, "a0", "원문A")])
    with pytest.raises(HTTPException) as gone_json:
        rw.get_paid_result_web("tok-1")
    with pytest.raises(HTTPException) as gone_xlsx:
        rw.get_paid_result_excel("tok-1")
    assert gone_json.value.status_code == 410
    assert gone_xlsx.value.status_code == 410

    rec_free = stored_rec([leg_obligation("a0", "점검")], tier="BUILDING_FREE")
    install(monkeypatch, rec_free, product_items=[source_item(0, "a0", "원문A")])
    free_json = rw.get_paid_result_web("tok-1")
    assert free_json["data"]["is_free"] is True
    assert "premium_result_v1" not in free_json["data"]
    with pytest.raises(HTTPException) as free_xlsx:
        rw.get_paid_result_excel("tok-1")
    assert free_xlsx.value.status_code == 403

    rec_legacy = stored_rec(
        [leg_obligation("a0", "점검")],
        tier="BUILDING_V2",
        rules_table=[{"law_name": "L", "law_article": "1", "obligation_summary": "x",
                      "description": "x", "rule_id": "r1"}],
    )
    monkeypatch.setattr(rw, "enrich_rules_with_candidate_slots", lambda *a, **k: None)
    install(monkeypatch, rec_legacy, product_items=[source_item(0, "a0", "원문A")])
    legacy_json = rw.get_paid_result_web("tok-1")
    assert "premium_result_v1" not in legacy_json["data"]
    with pytest.raises(HTTPException) as legacy_xlsx:
        rw.get_paid_result_excel("tok-1")
    assert legacy_xlsx.value.status_code == 404


def test_summary_skips_missing_does_not_invent_grade():
    p = _premium()
    p["profile"]["workers"] = None
    p["profile"]["floor_area"] = None
    wb = _load(build_paid_result_excel_v1(p))
    fields = [r[0] for r in _rows(wb["Summary"])]
    values = [r[1] for r in _rows(wb["Summary"])]
    assert "workers" not in fields
    assert "floor_area" not in fields
    dumped = "\n".join(str(v) for v in values)
    assert "소규모" not in dumped
    assert "등급" not in dumped
    wb.close()


def test_filename_uses_stored_values_only():
    p = _premium()
    name = excel_filename(p)
    assert name.startswith("TAI_법령진단_샘플 건설현장_")
    assert name.endswith(".xlsx")
    assert "datetime" not in name
    empty = _premium()
    empty["profile"]["company_name"] = None
    assert excel_filename(empty) == "TAI_법령진단_결과.xlsx"
    assert "attachment;" in content_disposition(name)


def test_schedule_and_assignments_are_1to1():
    p = _premium()
    wb = _load(build_paid_result_excel_v1(p))
    assert len(_rows(wb["Schedule"])) == 2
    assert len(_rows(wb["Assignments"])) == 2
    assert _header(wb["Schedule"]) == [
        "ref", "law_name", "law_article", "when", "inspection_cycle", "raw_cycle", "conflict",
    ]
    dumped = "\n".join("" if v is None else str(v) for v in _cell_values(wb))
    for banned in ("예정일", "D-day", "다음점검일", "overdue", "안전관리자 배정", "위탁가능"):
        assert banned not in dumped
    wb.close()


def test_local_xlsx_smoke_portfolio(tmp_path):
    """실제 .xlsx 1개 생성(temp). repo commit 0."""
    obs = [_ob(i, article="375" if i in (17, 18) else str(i), how="@foo" if i == 0 else None)
           for i in range(23)]
    sources = [{"ref": i, "text": "원문-{}".format(i)} for i in range(23)]
    articles = []
    for i in range(8):
        articles.append({
            "law_name": "산업안전보건기준에 관한 규칙",
            "article_no": 100 + i,
            "article_sub_no": None,
            "article_title": "조문{}".format(i),
            "article_text": "조문{} 원문\n줄바꿈 유지".format(i),
            "related_refs": [i],
        })
    articles.append({
        "law_name": "산업안전보건기준에 관한 규칙",
        "article_no": 375,
        "article_sub_no": None,
        "article_title": "화물취급",
        "article_text": "제375조 " + ("본문" * 50),
        "related_refs": [17, 18],
    })
    p = _premium()
    p["materials"]["obligations"] = obs
    p["materials"]["overview"]["total_obligation_count"] = 23
    p["canonical_sources"] = sources
    p["evidence"]["articles"] = articles
    raw = build_paid_result_excel_v1(p)
    sha = hashlib.sha256(raw).hexdigest()
    path = tmp_path / "premium_smoke.xlsx"
    path.write_bytes(raw)
    wb = _load(raw)
    formula_count = 0
    truncated = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    formula_count += 1
                if isinstance(cell.value, str) and len(cell.value) > EXCEL_CELL_MAX:
                    truncated += 1
    canon_col = _header(wb["Obligations"]).index("canonical_source_text")
    exact = sum(1 for r in _rows(wb["Obligations"]) if r[canon_col] and r[canon_col] != CANONICAL_UNAVAILABLE)
    stats = {
        "file_size": len(raw),
        "sha256": sha,
        "sheet_names": list(wb.sheetnames),
        "sheet_row_counts": {n: wb[n].max_row for n in wb.sheetnames},
        "obligation_count": len(_rows(wb["Obligations"])),
        "canonical_exact_count": exact,
        "evidence_article_count": len(articles),
        "formula_count": formula_count,
        "truncated_text_count": truncated,
    }
    (tmp_path / "smoke_stats.json").write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")
    assert stats["obligation_count"] == 23
    assert stats["formula_count"] == 0
    assert stats["truncated_text_count"] == 0
    assert stats["file_size"] > 0
    wb.close()
    # 환경 변수로 리포트 경로를 남긴다(커밋하지 않음).
    os.environ["TAI_EXCEL_SMOKE_SHA256"] = sha
    os.environ["TAI_EXCEL_SMOKE_SIZE"] = str(len(raw))
    print("LOCAL_XLSX_SMOKE", json.dumps(stats, ensure_ascii=False))
